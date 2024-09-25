from django.shortcuts import render, redirect
from .models import *
from .forms import *
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from datetime import datetime
from django.http import JsonResponse, HttpResponse
import decimal
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.units import mm
from reportlab.lib import colors
import openpyxl
from openpyxl.utils import get_column_letter

# Create your views here.
@login_required
def dashboard(request):
    mcaja = MovimentoCaja.objects.all().count()
    ventas = Venta.objects.all().count()
    context = {
        'mcaja': mcaja,
        'ventas': ventas,
    }
    return render(request, 'index.html', context)

def signout(request):
    logout(request)
    return redirect('login')

def sigin(request):
    if request.method == 'POST':
        user = authenticate(
            request, username=request.POST['username'], password=request.POST['password'])
        login(request, user)
        return redirect('inicio')
    return render(request, 'login.html')

def registro(request):
    e_form = EmployeForm(request.POST or None)
    c_form = ContactoForm(request.POST or None)
    r_form = UserForm(request.POST or None)
    context = {
        'e_form': e_form,
        'c_form': c_form,
        'r_form': r_form,
    }

    if request.method == 'POST':
        try:
            if e_form.is_valid() and c_form.is_valid() and r_form.is_valid():
                e = e_form.save(commit=False)
                c = c_form.save()
                r = r_form.save()
                e.contacto = c
                e.usuario = r
                e.save()
                login(request, r)
                return redirect('inicio')
            else:
                print(e_form.errors)
                print(c_form.errors)
                print(r_form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'registro.html', context)

@login_required
def negocio(request):
    negocios = Negocio.objects.all()

    context = {'negocios': negocios}

    return render(request, 'negocio/list.html', context)

@login_required
def negocio_crear(request):
    n_form = NegocioForm(request.POST or None)
    c_form = ContactoForm(request.POST or None)
    d_form = DomicilioForm(request.POST or None)
    f_form = FiscalForm(request.POST or None)

    context = {
        'n_form': n_form,
        'c_form': c_form,
        'd_form': d_form,
        'f_form': f_form
    }

    if request.method == 'POST':
        try:
            if n_form.is_valid() and c_form.is_valid() and d_form.is_valid() and f_form.is_valid():
                n = n_form.save(commit=False)
                f = f_form.save(commit=False)
                d = d_form.save()
                c = c_form.save()
                f.domicilio = d
                f.save()
                n.contacto = c
                n.fiscal = f
                n.save()
                return redirect('negocio')
            else:
                print(n_form.errors)
                print(c_form.errors)
                print(d_form.errors)
                print(f_form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'negocio/create.html', context)

@login_required
def negocio_editar(request, pk):
    negocio = Negocio.objects.get(pk=pk)
    contacto = negocio.contacto
    fiscal = negocio.fiscal
    domicilio = negocio.fiscal.domicilio
    n_form = NegocioForm(request.POST or None, instance=negocio)
    c_form = ContactoForm(request.POST or None, instance=contacto)
    d_form = DomicilioForm(request.POST or None, instance=domicilio)
    f_form = FiscalForm(request.POST or None, instance=fiscal)

    context = {
        'n_form': n_form,
        'c_form': c_form,
        'd_form': d_form,
        'f_form': f_form
    }

    if request.method == 'POST':
        try:
            if n_form.is_valid() and c_form.is_valid() and d_form.is_valid() and f_form.is_valid():
                d_form.save()
                f_form.save()
                c_form.save()
                n_form.save()
                return redirect('negocio')
            else:
                print(n_form.errors)
                print(c_form.errors)
                print(d_form.errors)
                print(f_form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'negocio/edit.html', context)

@login_required
def negocio_eliminar(request, pk):
    negocio = Negocio.objects.get(pk=pk)
    contacto = negocio.contacto
    fiscal = negocio.fiscal
    domicilio = negocio.fiscal.domicilio

    if request.method == 'POST':
        domicilio.delete()
        fiscal.delete()
        contacto.delete()
        negocio.delete()
        return redirect('negocio')

@login_required    
def marca(request):
    marcas = Marca.objects.all()
    context = {'marcas': marcas}

    return render(request, 'marca/list.html', context)

@login_required
def marca_crear(request):
    form = MarcaForm(request.POST or None)
    context = {'form':form}

    if request.method == 'POST':
        try:
            if form.is_valid():
                form.save()
                return redirect('marca')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'marca/create.html', context)

@login_required
def marca_editar(request, pk):
    marca = Marca.objects.get(pk=pk)
    form = MarcaForm(request.POST or None, instance=marca)
    context = {
        'form': form
    }

    if request.method == 'POST':
        try:
            if form.is_valid():
                form.save()
                return redirect('marca')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'marca/edit.html', context)

@login_required
def marca_eliminar(request, pk):
    marca = Marca.objects.get(pk=pk)

    if request.method == 'POST':
        marca.delete()
        return redirect('marca')

@login_required    
def cliente(request):
    clientes = Cliente.objects.all()
    context = {'clientes': clientes}

    return render(request, 'cliente/list.html', context)

@login_required
def cliente_crear(request):
    c_form = ClienteForm(request.POST or None)
    con_form = ContactoForm(request.POST or None)
    f_form = FiscalForm(request.POST or None)
    d_form = DomicilioForm(request.POST or None)
    context = {
        'c_form': c_form,
        'con_form': con_form,
        'f_form': f_form,
        'd_form': d_form
    }

    if request.method == 'POST':
        try:
            if c_form.is_valid() and con_form.is_valid() and f_form.is_valid() and d_form.is_valid():
                cli = c_form.save(commit=False)
                f = f_form.save(commit=False)
                d = d_form.save()
                f.domicilio = d
                f.save()
                con = con_form.save()
                cli.contacto = con
                cli.fiscal = f
                cli.save()
                return redirect('cliente')
            else:
                print(c_form.errors)
                print(con_form.errors)
                print(f_form.errors)
                print(d_form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'cliente/create.html', context)

@login_required
def cliente_editar(request, pk):
    cliente = Cliente.objects.get(pk=pk)
    contacto = cliente.contacto
    fiscal = cliente.fiscal
    domiclio = cliente.fiscal.domicilio
    c_form = ClienteForm(request.POST or None, instance=cliente)
    con_form = ContactoForm(request.POST or None, instance=contacto)
    f_form = FiscalForm(request.POST or None, instance=fiscal)
    d_form = DomicilioForm(request.POST or None, instance=domiclio)
    context = {
        'c_form': c_form,
        'con_form': con_form,
        'f_form': f_form,
        'd_form': d_form
    }

    if request.method == 'POST':
        try:
            if c_form.is_valid() and con_form.is_valid() and f_form.is_valid() and d_form.is_valid():
                con_form.save()
                d_form.save()
                f_form.save()
                c_form.save()
                return redirect('cliente')
            else:
                print(c_form.errors)
                print(con_form.errors)
                print(f_form.errors)
                print(d_form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'cliente/edit.html', context)

@login_required
def cliente_eliminar(request, pk):
    cliente = Cliente.objects.get(pk=pk)
    contacto = cliente.contacto
    fiscal = cliente.fiscal
    domiclio = cliente.fiscal.domicilio

    if request.method == 'POST':
        contacto.delete()
        domiclio.delete()
        fiscal.delete()
        cliente.delete()
        return redirect('cliente')

@login_required    
def caja(request):
    cajas = Caja.objects.all()
    empleado = Empleado.objects.get(usuario=request.user)
    context = {
        'cajas':cajas,
        'empleado': empleado
        }

    return render(request, 'caja/list.html', context)

@login_required
def caja_agregar(request):
    form = CajaForm(request.POST or None)
    context = {'form':form}

    if request.method == 'POST':
        try:
            if form.is_valid():
                form.save()
                return redirect('caja')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'caja/create.html', context)

@login_required
def caja_editar(request, pk):
    caja = Caja.objects.get(pk=pk)
    form = CajaForm(request.POST or None, instance=caja)
    context = {'form':form}

    if request.method == 'POST':
        try:
            if form.is_valid():
                form.save()
                return redirect('caja')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'caja/edit.html', context)

@login_required
def caja_eliminar(request, pk):
    caja = Caja.objects.get(pk=pk)

    if request.method == 'POST':
        caja.delete()
        return redirect('caja')

@login_required    
def departamento(request):
    departamantos = Departamento.objects.all()
    context = {'departamentos': departamantos}

    return render(request, 'departamento/list.html', context)

@login_required
def departamento_crear(request):
    form = DepartamanetoForm(request.POST or None)
    context = {'form': form}

    if request.method == 'POST':
        try:
            if form.is_valid():
                form.save()
                return redirect('departamento')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'departamento/create.html', context)

@login_required
def departamento_editar(request, pk):
    departamento = Departamento.objects.get(pk=pk)
    form = DepartamanetoForm(request.POST or None, instance=departamento)
    context = {'form': form}

    if request.method == 'POST':
        try:
            if form.is_valid():
                form.save()
                return redirect('departamento')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'departamento/edit.html', context)

@login_required
def departamento_eliminar(request, pk):
    departamento = Departamento.objects.get(pk=pk)

    if request.method == 'POST':
        departamento.delete()
        return redirect('departamento')

@login_required    
def medida(request):
    medidas = Medida.objects.all()
    context = {'medidas': medidas}

    return render(request, 'medida/list.html', context)

@login_required
def medida_crear(request):
    form = MedidaForm(request.POST or None)
    context = {'form': form}

    if request.method == 'POST':
        try:
            if form.is_valid():
                form.save()
                return redirect('medida')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'medida/create.html', context)

@login_required
def medida_editar(request, pk):
    medida = Medida.objects.get(pk=pk)
    form = MedidaForm(request.POST or None, instance=medida)
    context = {'form': form}

    if request.method == 'POST':
        try:
            if form.is_valid():
                form.save()
                return redirect('medida')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'medida/edit.html', context)

@login_required
def medida_eliminar(request, pk):
    medida = Medida.objects.get(pk=pk)

    if request.method == 'POST':
        medida.delete()
        return redirect('medida')

@login_required    
def producto(request):
    productos = Producto.objects.all()
    context = {'productos': productos}

    return render(request, 'producto/list.html', context)

@login_required
def producto_crear(request):
    form = ProductoForm(request.POST or None)
    context = {'form': form}

    if request.method == 'POST':
        try:
            if form.is_valid():
                form.save()
                return redirect('producto')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'producto/create.html', context)

@login_required
def producto_editar(request, pk):
    producto = Producto.objects.get(pk=pk)
    form = ProductoForm(request.POST or None, instance=producto)
    context = {'form': form}

    if request.method == 'POST':
        try:
            if form.is_valid():
                form.save()
                return redirect('producto')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'producto/edit.html', context)

@login_required
def producto_eliminar(request, pk):
    producto = Producto.objects.get(pk=pk)

    if request.method == 'POST':
        producto.delete()
        return redirect('producto')

@login_required   
def empleado(request):
    empleados = Empleado.objects.all()
    context = {'empleados': empleados}

    return render(request, 'empleado/list.html', context)

@login_required
def empleado_crear(request):
    e_form = EmpleadoForm(request.POST or None)
    u_form = UserForm(request.POST or None)
    c_form = ContactoForm(request.POST or None)
    context = {
        'e_form': e_form,
        'u_form': u_form,
        'c_form': c_form
    }

    if request.method == 'POST':
        try:
            if e_form.is_valid() and u_form.is_valid() and c_form.is_valid():
                e = e_form.save(commit=False)
                u = u_form.save()
                c = c_form.save()
                e.usuario = u
                e.contacto = c
                e.save()
                return redirect('empleado')
            else:
                print(e_form.errors)
                print(u_form.errors)
                print(c_form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'empleado/create.html', context)

@login_required
def empleado_editar(request, pk):
    empleado = Empleado.objects.get(pk=pk)
    usuario = empleado.usuario
    contacto = empleado.contacto
    e_form = EmpleadoForm(request.POST or None, instance=empleado)
    u_form = UserForm(request.POST or None, instance=usuario)
    c_form = ContactoForm(request.POST or None, instance=contacto)
    context = {
        'e_form': e_form,
        'u_form': u_form,
        'c_form': c_form
    }

    if request.method == 'POST':
        try:
            if e_form.is_valid() and u_form.is_valid() and c_form.is_valid():
                u_form.save()
                c_form.save()
                e_form.save()
                return redirect('empleado')
            else:
                print(e_form.errors)
                print(u_form.errors)
                print(c_form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'empleado/edit.html', context)

@login_required
def empleado_eliminar(request, pk):
    empleado = Empleado.objects.get(pk=pk)
    usuario = empleado.usuario
    contacto = empleado.contacto

    if request.method == 'POST':
        contacto.delete()
        usuario.delete()
        empleado.delete()
        return redirect('empleado')
    
@login_required
def abrir_caja(request, pk):
    caja = Caja.objects.get(pk=pk)
    empleado = Empleado.objects.get(usuario=request.user)
    form = MCajaForm(request.POST or None)
    context = {'form': form}
    request.session['caja_abierta'] = 0

    if request.method == 'POST':
        try:
            if form.is_valid():
                f = form.save(commit=False)
                f.caja = caja
                f.empleado = empleado
                f.fecha = datetime.now()
                f.movimiento = 'A'
                request.session['caja_abierta'] = caja.id
                f.save()
                return redirect('caja')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'caja/abrir.html', context)

@login_required
def cerrar_caja(request, pk):
    caja = Caja.objects.get(pk=request.session['caja_abierta'])
    caja_abierta = MovimentoCaja.objects.get(caja=caja, movimiento='A')
    empleado = Empleado.objects.get(usuario=request.user)
    form = MCajaForm(request.POST or None)
    context = {'form': form}

    if request.method == 'POST':
        try:
            if form.is_valid():
                f = form.save(commit=False)
                f.caja = caja
                f.empleado = empleado
                f.fecha = datetime.now()
                f.movimiento = 'C'
                f.monto_abierto = caja_abierta.monto_abierto
                f.saldo_final = f.monto_cierre - f.monto_abierto
                request.session['caja_abierta'] = 0
                f.save()
                return redirect('caja')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'caja/cerrar.html', context)

@login_required
def venta(request):
    ventas = Venta.objects.all()
    empleado = Empleado.objects.get(usuario=request.user)
    context = {'ventas': ventas, 'empleado': empleado}
    return render(request, 'venta/list.html', context)

@login_required
def venta_crear(request):
    form = VentaForm(request.POST or None)
    empleado = Empleado.objects.get(usuario=request.user)
    context = {'form': form}
    fecha = datetime.now()
    fecha_formato = fecha.strftime("%d/%m/%Y %H:%M:%S")

    if request.method == 'POST':
        try:
            if form.is_valid():
                f = form.save(commit=False)
                f.fecha = fecha_formato
                f.caja_id = request.session['caja_abierta']
                f.vendedor = empleado
                f.subtotal = f.cantidad * f.producto.precio
                if f.descuento is not None and f.descuento > 0:
                    descuento_decimal = decimal.Decimal(f.descuento) / 100
                    f.total = f.subtotal - (f.subtotal * descuento_decimal)
                else:
                    f.total = f.cantidad * f.producto.precio
                f.save()
                return redirect('venta')
            else:
                print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'venta/create.html', context)

def autocomplete_producto(request):
    if "term" in request.GET:
        qs = Producto.objects.filter(nombre__icontains=request.GET.get("term"))
        titles = list()
        for p in qs:
            datos = {
                'id': p.id,
                'label': p.nombre,
                'precio': p.precio,
                'existencia': p.existencia
            }
            titles.append(datos)
        return JsonResponse(titles, safe=False)

def autocomplete_cliente(request):
    if 'term' in request.GET:
        clientes = Cliente.objects.filter(nombre__icontains=request.GET.get('term'))
        titles = list()
        for cli in clientes:
            datos = {'id': cli.id,
                     'label': cli.nombre}
            titles.append(datos)
        return JsonResponse(titles, safe=False)

def sidebar(request):
    empleado = Empleado.objects.get(usuario=request.user)
    context = {'empleado': empleado}
    return render(request, 'sidebar.html', context)

@login_required
def venta_detalle(request, pk):
    venta = Venta.objects.get(pk=pk)
    context = {'venta': venta}

    return render(request, 'venta/detalle.html', context)

@login_required
def venta_pdf(request, pk):
    venta = Venta.objects.get(pk=pk)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="venta{venta.id}.pdf"'

    pdf = canvas.Canvas(response, pagesize=letter)

    # Cabecera
    pdf.setFont("Helvetica", 12)
    pdf.drawString(50, 750, "Detalle de la Venta")
    pdf.drawString(50, 730, f"Fecha: {venta.fecha}")
    pdf.drawString(350, 730, f"Vendedor: {venta.vendedor.nombre}")
    pdf.drawString(50, 710, f"Caja: {venta.caja.nombre}")
    pdf.drawString(350, 710, f"Cliente: {venta.cliente.nombre}")

    # Detalles del producto
    pdf.setFont("Helvetica", 10)
    pdf.drawString(50, 690, f"Producto: {venta.producto.nombre}")
    pdf.drawString(200, 690, f"Precio: ${venta.producto.precio}")
    pdf.drawString(50, 670, f"Cantidad: {venta.cantidad}")
    if not venta.descuento:
        pdf.drawString(50, 650, f"Descuento: 0%")
    else:
        pdf.drawString(50, 650, f"Descuento: {venta.descuento}%")


    # Subtotal
    pdf.drawString(50, 600, f"Subtotal: ${venta.subtotal}")

    # Línea divisoria
    pdf.line(50, 580, 550, 580)

    # Total
    pdf.drawString(50, 560, f"Total: ${venta.total}")

    pdf.showPage()
    pdf.save()
    return response
    
@login_required
def venta_ticket(request, pk):
    venta = Venta.objects.get(pk=pk)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ticket_{venta.id}.pdf"'
    
    # Crear un canvas de ReportLab para el PDF
    c = canvas.Canvas(response, pagesize=(80 * mm, 200 * mm))  # Tamaño aproximado de un ticket de compra

    # Establecer algunas variables para la posición del texto
    y_position = 190 * mm  # Posición inicial en Y
    x_margin = 5 * mm  # Margen izquierdo

    # Establecer el título del ticket
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x_margin, y_position, "TICKET DE COMPRA")
    
    # Dibujar una línea divisoria
    y_position -= 10
    c.setStrokeColor(colors.black)
    c.line(x_margin, y_position, 75 * mm, y_position)
    
    # Información de la venta
    y_position -= 10
    c.setFont("Helvetica", 8)
    c.drawString(x_margin, y_position, f"Fecha: {venta.fecha.strftime('%Y-%m-%d %H:%M:%S')}")
    
    y_position -= 10
    c.drawString(x_margin, y_position, f"Producto: {venta.producto.nombre}")
    
    y_position -= 10
    c.drawString(x_margin, y_position, f"Precio: ${venta.producto.precio:.2f}")
    
    y_position -= 10
    c.drawString(x_margin, y_position, f"Cantidad: {venta.cantidad}")
    
    if venta.descuento:
        y_position -= 10
        c.drawString(x_margin, y_position, f"Descuento: {venta.descuento}%")
    
    y_position -= 10
    c.drawString(x_margin, y_position, f"Subtotal: ${venta.subtotal:.2f}")
    
    y_position -= 10
    c.setFont("Helvetica-Bold", 8)
    c.drawString(x_margin, y_position, f"TOTAL: ${venta.total:.2f}")

    # Información final (ej. agradecimiento)
    y_position -= 20
    c.setFont("Helvetica", 8)
    c.drawString(x_margin, y_position, "Gracias por su compra")

    # Finalizar el PDF
    c.showPage()
    c.save()

    return response

@login_required
def venta_excel(request):
     # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Encabezados de las columnas
    encabezados = ['Fecha', 'Subtotal', 'Total', 'Cantidad', 'Vendedor', 'Caja', 'Producto', 'Descuento', 'Cliente']
    
    # Escribir los encabezados en la primera fila
    for col_num, encabezado in enumerate(encabezados, 1):
        col_letra = get_column_letter(col_num)
        ws[f'{col_letra}1'] = encabezado

    # Obtener todos los registros de Venta
    ventas = Venta.objects.all()

    # Escribir los datos de las ventas
    for row_num, venta in enumerate(ventas, 2):
        ws[f'A{row_num}'] = venta.fecha.strftime('%Y-%m-%d %H:%M:%S')
        ws[f'B{row_num}'] = float(venta.subtotal)
        ws[f'C{row_num}'] = float(venta.total)
        ws[f'D{row_num}'] = venta.cantidad
        ws[f'E{row_num}'] = str(venta.vendedor.nombre) if venta.vendedor else 'N/A'
        ws[f'F{row_num}'] = str(venta.caja.nombre) if venta.caja else 'N/A'
        ws[f'G{row_num}'] = str(venta.producto.nombre) if venta.producto else 'N/A'
        ws[f'H{row_num}'] = venta.descuento if venta.descuento is not None else '0'
        ws[f'I{row_num}'] = str(venta.cliente.nombre) if venta.cliente else 'N/A'

    # Generar la respuesta HTTP con el archivo adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'
    wb.save(response)
    
    return response
